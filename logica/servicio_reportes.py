import openpyxl
from datetime import datetime
import os

def registrar_venta_en_excel(folio, subtotal, descuento, total, nombre_cajero, metodo_pago, carrito):
    ruta = "Reporte_Ventas_Boutique.xlsx"
    
    try:
        if not os.path.exists(ruta):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registro de Ventas"
            ws.append([
                "Folio", "Fecha y Hora", "Cajero", "Método de Pago",
                "Subtotal ($)", "Descuento ($)", "Total ($)", "Detalle de Productos"
            ])
        else:
            wb = openpyxl.load_workbook(ruta)
            ws = wb.active
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            expected = [
                "Folio", "Fecha y Hora", "Cajero", "Método de Pago",
                "Subtotal ($)", "Descuento ($)", "Total ($)", "Detalle de Productos"
            ]
            if header[:len(expected)] != expected:
                for col_index, heading in enumerate(expected, start=1):
                    ws.cell(row=1, column=col_index, value=heading)

        detalle_items = {}
        for item in carrito:
            if item.id not in detalle_items:
                detalle_items[item.id] = {
                    "nombre": item.nombre,
                    "precio": item.precio,
                    "cantidad": 0,
                }
            detalle_items[item.id]["cantidad"] += 1

        lineas_detalle = []
        for info in detalle_items.values():
            subtotal_item = info["cantidad"] * info["precio"]
            lineas_detalle.append(
                f"{info['cantidad']}x {info['nombre']} @ ${info['precio']:.2f} = ${subtotal_item:.2f}"
            )

        detalle_texto = "\n".join(lineas_detalle)
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws.append([folio, fecha, nombre_cajero, metodo_pago, subtotal, descuento, total, detalle_texto])

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 60

        wb.save(ruta)
        print(" 📊 Registro en Excel actualizado.")
        return True

    except PermissionError:
        print("\n ⚠️ ¡ERROR DE PERMISO!")
        print(" No se pudo actualizar el Excel porque el archivo está ABIERTO.")
        print(" Por favor, ciérrelo para que las próximas ventas se registren correctamente.")
        return False

    except Exception as e:
        print(f" ❌ Error inesperado en Excel: {e}")
        return False


def generar_ticket_txt(folio, subtotal, descuento, total, nombre_cajero, metodo_pago, efectivo, carrito):
    if not os.path.exists("facturas"):
        os.makedirs("facturas")
        
    ruta_ticket = f"facturas/ticket_{folio}.txt"
    
    resumen_dict = {}
    for item in carrito:
        if item.id not in resumen_dict:
            resumen_dict[item.id] = {"nombre": item.nombre, "precio": item.precio, "cant": 0}
        resumen_dict[item.id]["cant"] += 1

    try:
        with open(ruta_ticket, "w", encoding="utf-8") as f:
            f.write("================================\n")
            f.write("        POS BOUTIQUE ZAMORA      \n")
            f.write("================================\n")
            f.write(f"Folio: #{folio}\n")
            f.write(f"Cajero: {nombre_cajero}\n")
            f.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
            f.write(f"Pago: {metodo_pago}\n")
            f.write("--------------------------------\n")

            f.write(f"{'Cant':<4} {'Producto':<18} {'Unit':>7} {'Total':>8}\n")
            f.write("--------------------------------\n")
            for info in resumen_dict.values():
                subtotal_item = info['cant'] * info['precio']
                nombre_corto = info['nombre'][:18]
                f.write(f"{info['cant']:<4} {nombre_corto:<18} ${info['precio']:>6.2f} ${subtotal_item:>7.2f}\n")

            f.write("--------------------------------\n")
            f.write(f"SUBTOTAL:           ${subtotal:>8.2f}\n")
            if descuento > 0:
                f.write(f"DESCUENTO:          -${descuento:>7.2f}\n")
            f.write(f"TOTAL:              ${total:>8.2f}\n")
            if metodo_pago == "Efectivo":
                f.write(f"EFECTIVO:           ${efectivo:>8.2f}\n")
                cambio = efectivo - total
                if cambio < 0:
                    cambio = 0.0
                f.write(f"CAMBIO:             ${cambio:>8.2f}\n")
            f.write("================================\n")
            f.write("      ¡Gracias por su compra!     \n")
        print(f" 🧾 Ticket generado: {ruta_ticket}")
    except Exception as e:
        print(f" ❌ Error al generar ticket TXT: {e}")


def registrar_devolucion_en_excel(nombre_cajero, producto):
    ruta = "Reporte_Ventas_Boutique.xlsx"
    if not os.path.exists(ruta):
        return False # Si no hay archivo, no hacemos nada
        
    try:
        wb = openpyxl.load_workbook(ruta)
        ws = wb.active
        
        fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
        # Registramos como DEV, monto negativo, y el detalle
        ws.append(["DEV", fecha, nombre_cajero, -producto.precio, f"DEVOLUCIÓN: 1x {producto.nombre}"])
        
        wb.save(ruta)
        return True
    except PermissionError:
        print("\n ⚠️ No se pudo registrar la devolución en Excel (Archivo abierto).")
        return False
    except Exception as e:
        print(f"\n ❌ Error en Excel: {e}")
        return False    